import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="Route Summary Master Pro", layout="wide")
st.title("🚌 Monthly Route Summary Automator")

# Master Route Code Mapping
ROUTE_MAPPING = {
    "GITHARI": "T65", "BRIDGES": "T04", "CENTRE": "T33",
    "CHEBA NGURUKA": "T05", "CHEBA-B": "T57", "CHOBE": "T14",
    "CHUMA": "T09", "CIONDO-B": "T13", "DAM": "T18",
    "EXLEES": "T68", "FARU": "T16", "GACATA": "T40",
    "GACHUCHA-B": "T46", "GATHARA": "T67", "GATITU": "T43",
    "GITHIMA": "T29", "GITIRI": "T31", "GITITE": "T12",
    "GURD": "T52", "KIBORE": "T41", "KIRIMA": "T61",
    "MAIN": "T87", "SEMIHEADQUATER": "T59", "THINDI": "T20",
    "UPPER CIONDO": "T38", "WANGU": "T82", "YAANGA": "T27",
    "KWARE": "T90", "KARIMA": "T91", "MUTAMAIYU": "T39",
    "MUTAMAIYU PM": "T72", "CHUMA-B": "T95"
}

def style_excel_workbook(writer, df_master, existing_sheets):
    """Applies professional styling, auto-adjusts column widths, and formats numbers."""
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    bold_font = Font(name="Calibri", size=11, bold=True)
    
    # Style Monthly Summary
    ws_master = writer.sheets["Monthly Summary"]
    for col_idx in range(1, len(df_master.columns) + 1):
        cell = ws_master.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx in range(2, len(df_master) + 2):
        is_total_row = (row_idx == len(df_master) + 1)
        for col_idx in range(1, len(df_master.columns) + 1):
            cell = ws_master.cell(row=row_idx, column=col_idx)
            if is_total_row:
                cell.fill = total_fill
                cell.font = bold_font
            if col_idx in [2, 3, 4, 5, 6]:
                cell.number_format = '#,##0.0'
            elif col_idx == 7:
                cell.number_format = 'KSh #,##0.00'

    # Auto-adjust column widths across all sheets & style total rows/columns
    for sheetname in writer.sheets:
        ws_curr = writer.sheets[sheetname]
        
        # Highlight total row in daily breakdown sheets
        if "Daily" in sheetname:
            max_r = ws_curr.max_row
            max_c = ws_curr.max_column
            for c in range(1, max_c + 1):
                ws_curr.cell(row=max_r, column=c).fill = total_fill
                ws_curr.cell(row=max_r, column=c).font = bold_font
            for r in range(1, max_r + 1):
                ws_curr.cell(row=r, column=max_c).font = bold_font

        for col in ws_curr.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_curr.column_dimensions[col_letter].width = max(max_len + 4, 12)

# --- APP LAYOUT ---
col1, col2 = st.columns([1, 1])

with col1:
    master_file = st.file_uploader("📁 Step 1: Upload Master File (Optional for Week 1)", type=["xlsx", "xls"], key="master")

with col2:
    selected_week = st.selectbox("📅 Step 2: Select Week to Add/Update:", ["week 1", "week 2", "week 3", "week 4"])
    daily_file = st.file_uploader(f"📄 Step 3: Upload Raw Daily File for {selected_week}", type=["xlsx", "xls"], key="daily")
    rate_multiplier = st.number_input("Rate Multiplier per Unit (KSh):", value=60.0, step=1.0)

if daily_file is not None:
    if st.button("🚀 Process & Update Master Workbook", type="primary"):
        try:
            xls_daily = pd.ExcelFile(daily_file)
            df_raw = pd.read_excel(xls_daily, sheet_name=0)
            
            day_cols_info = []
            header_has_days = any(str(c).strip().isdigit() for c in df_raw.columns)
            
            if header_has_days:
                for c_idx, col_name in enumerate(df_raw.columns):
                    c_str = str(col_name).strip()
                    if c_str.isdigit():
                        day_cols_info.append((f"Day {int(c_str)}", c_idx - 1, c_idx))
                start_row = 1
            else:
                for c_idx in range(len(df_raw.columns)):
                    val = df_raw.iloc[0, c_idx]
                    if pd.notna(val) and str(val).strip().isdigit():
                        day_cols_info.append((f"Day {int(str(val).strip())}", c_idx - 1, c_idx))
                start_row = 2

            route_records = {}
            unmapped_routes = set()
            for day_label, route_col, val_col in day_cols_info:
                for r in range(start_row, len(df_raw)):
                    r_name = df_raw.iloc[r, route_col]
                    r_var = df_raw.iloc[r, val_col]
                    if pd.notna(r_name) and str(r_name).strip() not in ["", "0", "Route", "nan", "NaN"]:
                        r_clean = str(r_name).strip().upper()
                        if r_clean not in ROUTE_MAPPING:
                            unmapped_routes.add(r_clean)
                        if r_clean not in route_records:
                            route_records[r_clean] = {}
                        v_num = pd.to_numeric(r_var, errors='coerce') if pd.notna(r_var) else 0.0
                        route_records[r_clean][day_label] = float(v_num) if pd.notna(v_num) else 0.0

            if unmapped_routes:
                st.warning(f"⚠️ Unmapped Routes Detected: {', '.join(unmapped_routes)}. Default code 'T00' applied.")

            # Construct Daily Grid
            df_daily_grid = pd.DataFrame.from_dict(route_records, orient='index').fillna(0)
            df_daily_grid.index.name = "Route"
            df_daily_grid = df_daily_grid.reset_index()
            day_cols_sorted = sorted([c for c in df_daily_grid.columns if c != "Route"], key=lambda x: int(x.replace("Day ", "")))
            
            # Sort routes and calculate weekly sum per route before adding total row/cols
            df_daily_grid = df_daily_grid[["Route"] + day_cols_sorted].sort_values("Route")
            weekly_totals = df_daily_grid.set_index("Route")[day_cols_sorted].sum(axis=1).to_dict()

            # Add Row Total Column ("Total")
            df_daily_grid["Total"] = df_daily_grid[day_cols_sorted].sum(axis=1)

            # Add Bottom Summary Total Row ("SUM TOTAL")
            daily_totals_row = {"Route": "SUM TOTAL"}
            for col in day_cols_sorted + ["Total"]:
                daily_totals_row[col] = df_daily_grid[col].sum()

            df_daily_grid_final = pd.concat([df_daily_grid, pd.DataFrame([daily_totals_row])], ignore_index=True)

            # --- Master Summary Update ---
            existing_sheets = {}
            if master_file is not None:
                xls_master = pd.ExcelFile(master_file)
                for s in xls_master.sheet_names:
                    existing_sheets[s] = pd.read_excel(xls_master, sheet_name=s)
                df_master = existing_sheets.get("Monthly Summary", existing_sheets[xls_master.sheet_names[0]])
            else:
                df_master = pd.DataFrame(columns=["Route", "week 1", "week 2", "week 3", "week 4", "sum", "AMOUNT", "Code"])

            if len(df_master) > 0 and df_master.iloc[0]["Route"] == "Route":
                df_master.columns = df_master.iloc[0].values
                df_master = df_master.iloc[1:].reset_index(drop=True)

            df_master = df_master[df_master["Route"].astype(str).str.strip().str.upper() != "SUM TOTAL"].copy()

            for col in ["week 1", "week 2", "week 3", "week 4"]:
                if col not in df_master.columns:
                    df_master[col] = 0.0
                df_master[col] = pd.to_numeric(df_master[col], errors='coerce').fillna(0.0)

            route_series = df_master["Route"].astype(str).str.strip().str.upper()
            all_routes = set(route_series).union(set(weekly_totals.keys()))
            master_dict = df_master.set_index(df_master["Route"].astype(str).str.strip().str.upper()).to_dict('index')
            
            updated_rows = []
            for r in sorted(all_routes):
                row_data = master_dict.get(r, {"Route": r, "week 1": 0.0, "week 2": 0.0, "week 3": 0.0, "week 4": 0.0})
                if r in weekly_totals:
                    row_data[selected_week] = weekly_totals[r]
                
                w_sum = sum([float(row_data.get(w, 0.0)) for w in ["week 1", "week 2", "week 3", "week 4"]])
                amt = w_sum * rate_multiplier
                code = ROUTE_MAPPING.get(r, "T00")
                
                row_data["Route"] = row_data.get("Route", r)
                row_data["sum"] = w_sum
                row_data["AMOUNT"] = amt
                row_data["Code"] = code
                updated_rows.append(row_data)

            df_updated_master = pd.DataFrame(updated_rows)[["Route", "week 1", "week 2", "week 3", "week 4", "sum", "AMOUNT", "Code"]]

            total_row = {
                "Route": "SUM TOTAL",
                "week 1": df_updated_master["week 1"].sum(),
                "week 2": df_updated_master["week 2"].sum(),
                "week 3": df_updated_master["week 3"].sum(),
                "week 4": df_updated_master["week 4"].sum(),
                "sum": df_updated_master["sum"].sum(),
                "AMOUNT": df_updated_master["AMOUNT"].sum(),
                "Code": ""
            }
            df_final_master = pd.concat([df_updated_master, pd.DataFrame([total_row])], ignore_index=True)

            existing_sheets["Monthly Summary"] = df_final_master
            existing_sheets[f"{selected_week.capitalize()} Daily"] = df_daily_grid_final

            st.success(f"Successfully processed **{selected_week}**!")

            # Key Metrics Display
            m1, m2, m3 = st.columns(3)
            m1.metric(f"Total Volume ({selected_week})", f"{df_updated_master[selected_week].sum():,.1f}")
            m2.metric("Cumulative Monthly Volume", f"{df_updated_master['sum'].sum():,.1f}")
            m3.metric("Cumulative Amount", f"KSh {df_updated_master['AMOUNT'].sum():,.2f}")

            tab1, tab2 = st.tabs(["📊 Cumulative Monthly Summary", f"📅 {selected_week.capitalize()} Daily Breakdown"])
            with tab1:
                st.dataframe(df_final_master, use_container_width=True)
            with tab2:
                st.dataframe(df_daily_grid_final, use_container_width=True)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final_master.to_excel(writer, sheet_name="Monthly Summary", index=False)
                for s_name, s_df in existing_sheets.items():
                    if s_name != "Monthly Summary":
                        s_df.to_excel(writer, sheet_name=s_name, index=False)
                
                style_excel_workbook(writer, df_final_master, existing_sheets)

            st.download_button(
                label="📥 Download Formatted Master Workbook (.xlsx)",
                data=output.getvalue(),
                file_name="Monthly_Route_Summary_Master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error processing workbook: {e}")
